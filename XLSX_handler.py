from traceback import print_exc

from PyQt5.QtWidgets import QFileDialog
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from loguru import logger
import models
from database import init_db, SESSIONLOCAL, ENGINE


def count_subelem(arr, idx):
    counter = 0
    for i in arr:
        if i[idx] == arr[0][idx]:
            counter += 1
    return counter


def divide_by(arr, idx):
    res = []
    sort_by(arr, idx)
    split_idx = count_subelem(arr, idx)
    a1 = arr[:split_idx]
    a2 = arr[split_idx:]
    try:
        if len(a1) != 0 and a1[2] < a2[2]:
            sort_by(a1, 2)
            res.append(a1)
            res.append(a2)
        elif len(a2) != 0 and a1[2] > a2[2]:
            sort_by(a2, 2)
            res.append(a2)
            res.append(a1)
        else:
            res.append(a1)
            res.append(a2)
    except:
        logger.info("Can't sort")
        print_exc()
        res.append(a1)
        res.append(a2)

    if [] in res:
        res.remove([])
    return res


def sort_by(arr, idx):
    n = len(arr)
    swapped = False
    for i in range(n - 1):
        for j in range(0, n - i - 1):
            if arr[j][idx] > arr[j + 1][idx]:
                swapped = True
                arr[j][idx], arr[j + 1][idx] = arr[j + 1][idx], arr[j][idx]
        if not swapped:
            break


def set_headers(ws, fields):
    red_fill = PatternFill(start_color='EB3B23',
                           end_color='EB3B23',
                           fill_type='solid')
    orange_fill = PatternFill(start_color='EB8433',
                              end_color='EB8433',
                              fill_type='solid')

    for i in range(len(fields)):
        ws.row_dimensions[1].height = 20
        ws.column_dimensions[get_column_letter(i * 3 + 1)].width = 15
        ws.column_dimensions[get_column_letter(i * 3 + 2)].width = 15
        ws.cell(row=1, column=i * 3 + 1).value = "Начало матча"
        ws.cell(row=1, column=i * 3 + 1).fill = red_fill
        ws.cell(row=1, column=i * 3 + 2).value = fields[i]
        ws.cell(row=1, column=i * 3 + 2).fill = orange_fill
        ws.cell(row=1, column=i * 3 + 3).fill = orange_fill


def load_to_file(database, file, games=None):
    #print("--"*20)
    logger.info("xuy 2")
    wb = Workbook()
    orange_fill = PatternFill(start_color='EB8433',
                              end_color='EB8433',
                              fill_type='solid')
    if not games:
        print("[picked from db]")
        games = database.query(models.Games).order_by(models.Games.stadium).all()
        res = []
        for game in games:
            res.append([game.team1, game.team2, game.start_time,
                        game.stadium, game.field, game.division])
    else:
        print("[without db]")
        res = games

    for elem in res:
        print(elem)
    sort_by(res, 2)
    res = divide_by(res, 3)

    for elem in res:
        ws = wb.create_sheet(elem[0][3])
        ws.title = elem[0][3]
        # ws.font = Font(bold=True)
        fields = list(set([i[4] for i in elem]))
        '''for i in range(2, len(fields) * 3 + 1, 2):
            for j in range(1, len(elem)+1):
                print(i, j)
                ws.cell(row=i, column=j).fill = gray_fill'''
        sort_by(elem, 2)
        set_headers(ws, fields)
        print(res.index(elem))
        for i in elem:
            print(i)
        for i in range(len(elem)):
            ws.cell(row=i + 2, column=fields.index(elem[i][4]) * 3 + 1).value = elem[i][2]
            ws.cell(row=i + 2, column=fields.index(elem[i][4]) * 3 + 2).value = (elem[i][0] + "\n" + elem[i][1])
            ws.cell(row=i + 2, column=fields.index(elem[i][4]) * 3 + 2).alignment = Alignment(wrapText=True)
            ws.cell(row=i + 2, column=fields.index(elem[i][4]) * 3 + 3).value = elem[i][-1]
            ws.cell(row=i + 2, column=fields.index(elem[i][4]) * 3 + 3).fill = orange_fill
            ws.row_dimensions[i + 2].height = 30
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(file)
    #database.metadata.drop_all(ENGINE)


if __name__ == "__main__":
    init_db()
    ENGINE.connect()
    database = SESSIONLOCAL()
    file = "C:\\Users\\Danil\\Desktop\\footballchik\\расписание.xlsx"
    load_to_file(database, file)
